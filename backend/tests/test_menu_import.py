import io
import zipfile

import openpyxl
from openpyxl import Workbook

from menu_import import ITEM_COLUMNS, GUIDE_FILENAME


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
def _wb_from_rows(rows, headers=None, sheet="Items"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers if headers is not None else ITEM_COLUMNS)
    for r in rows:
        ws.append(r)
    return wb


def _save(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _post(client, wb, headers):
    return client.post(
        "/api/admin/menu/import",
        data={"file": (_save(wb), "menu.xlsx")},
        content_type="multipart/form-data",
        headers=headers,
    )


def _xlsx_from_zip(raw):
    """Extract the template xlsx + zip info from a downloaded ZIP package.

    Returns (workbook, names, guide_text)."""
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        names = zf.namelist()
        wb = openpyxl.load_workbook(io.BytesIO(zf.read("menu-import-template.xlsx")))
        guide = zf.read(GUIDE_FILENAME).decode("utf-8")
    return wb, names, guide


# --------------------------------------------------------------------------
# Template download (ZIP: template + instruction guide)
# --------------------------------------------------------------------------
def test_template_download_is_zip_with_template_and_guide(client, auth_headers):
    resp = client.get("/api/admin/menu/template", headers=auth_headers)
    assert resp.status_code == 200
    assert resp.headers["Content-Type"] == "application/zip"
    assert "menu-import-template.zip" in resp.headers["Content-Disposition"]

    wb, names, guide = _xlsx_from_zip(resp.data)
    assert "menu-import-template.xlsx" in names
    assert GUIDE_FILENAME in names
    # the guide carries the key instructions
    assert "STEP-BY-STEP" in guide
    assert "AUTO-CREATION" in guide

    # the workbook itself still has both sheets + correct headers
    assert wb.sheetnames == ["Items", "Read me"]
    ws = wb["Items"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert headers == ITEM_COLUMNS


def test_template_requires_auth(client):
    assert client.get("/api/admin/menu/template").status_code == 401


def test_import_accepts_zip_package(client, db, auth_headers):
    # Download the real package, then add a row to the xlsx inside it and
    # re-upload the whole ZIP -> should import fine.
    pkg = client.get("/api/admin/menu/template", headers=auth_headers).data
    with zipfile.ZipFile(io.BytesIO(pkg)) as zf:
        wb = openpyxl.load_workbook(io.BytesIO(zf.read("menu-import-template.xlsx")))
    wb["Items"].append(["Masala Chai", "Spiced tea", 25, "Beverages", "Tea", "veg"])
    # re-zip the edited workbook
    buf = io.BytesIO()
    wb.save(buf)
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w") as zf:
        zf.writestr("menu-import-template.xlsx", buf.getvalue())
    zip_buf.seek(0)

    resp = client.post(
        "/api/admin/menu/import",
        data={"file": (zip_buf, "menu-import-template.zip")},
        content_type="multipart/form-data",
        headers=auth_headers,
    )
    assert resp.status_code == 200
    assert resp.get_json()["added"] == 1


# --------------------------------------------------------------------------
# Happy path
# --------------------------------------------------------------------------
def test_import_happy_path(client, db, auth_headers):
    wb = _wb_from_rows([
        ["Chicken Biryani", "Aromatic rice", 290, "Food", "Biryani", "non-veg",
         "Chicken", "Indian", "spicy", "heavy", "high", "yes", "no", ""],
        ["Masala Chai", "Spiced tea", 25, "Beverages", "Tea", "veg",
         "", "", "sweet", "light", "medium", "yes", "yes", "http://img"],
    ])
    resp = _post(client, wb, auth_headers)
    assert resp.status_code == 200
    body = resp.get_json()
    assert body["added"] == 2
    assert body["failed"] == 0
    assert body["total"] == 2
    assert body["added_names"] == ["Chicken Biryani", "Masala Chai"]

    items = db.read("restaurants/r1/items")
    assert len(items) == 2
    biryani = next(i for i in items.values() if i["name"] == "Chicken Biryani")
    assert biryani["price"] == 290
    assert biryani["item_type"] == "non-veg"
    assert biryani["main_ingredient"] == "Chicken"
    assert biryani["cuisine"] == "Indian"
    assert biryani["taste"] == ["spicy"]
    assert biryani["spice_level"] == 0
    # category links resolve to real nodes
    cats = db.read("restaurants/r1/categories")
    assert biryani["category_id"] in cats
    assert cats[biryani["category_id"]]["name"] == "Biryani"
    main_cats = db.read("restaurants/r1/main_categories")
    assert biryani["main_category_id"] in main_cats
    # vocab auto-created
    ings = db.read("restaurants/r1/ingredients")
    assert any(v["name"] == "Chicken" for v in ings.values())


# --------------------------------------------------------------------------
# Duplicate handling
# --------------------------------------------------------------------------
def test_import_duplicate_vs_existing(client, db, auth_headers):
    db.seed("restaurants/r1/items", {"i1": {"name": "X"}})
    wb = _wb_from_rows([
        ["X", "", 10, "Food", "S", "veg"],
        ["Y", "", 12, "Food", "S", "veg"],
    ])
    resp = _post(client, wb, auth_headers)
    assert resp.status_code == 200  # partial success
    body = resp.get_json()
    assert body["added"] == 1
    assert body["failed"] == 1
    assert body["failures"][0]["name"] == "X"
    assert body["failures"][0]["reason"] == "Duplicate name"
    assert body["failures"][0]["row"] == 2
    items = db.read("restaurants/r1/items")
    assert {i["name"] for i in items.values()} == {"X", "Y"}


def test_import_within_file_duplicate(client, db, auth_headers):
    wb = _wb_from_rows([
        ["Z", "", 10, "Food", "S", "veg"],
        ["Z", "", 10, "Food", "S", "veg"],
    ])
    body = _post(client, wb, auth_headers).get_json()
    assert body["added"] == 1
    assert body["failed"] == 1
    assert body["failures"][0]["reason"] == "Duplicate name"


def test_import_case_insensitive_dup(client, db, auth_headers):
    wb = _wb_from_rows([
        ["Biryani", "", 10, "Food", "S", "veg"],
        ["BIRYANI", "", 10, "Food", "S", "veg"],
    ])
    body = _post(client, wb, auth_headers).get_json()
    assert body["added"] == 1
    assert body["failed"] == 1


# --------------------------------------------------------------------------
# Failures + status codes
# --------------------------------------------------------------------------
def test_import_all_fail_returns_400_with_report(client, db, auth_headers):
    wb = _wb_from_rows([
        ["A", "", "free", "Food", "S", "veg"],   # invalid price
        ["B", "", "", "Food", "S", "veg"],        # missing price
    ])
    resp = _post(client, wb, auth_headers)
    assert resp.status_code == 400
    body = resp.get_json()
    assert body["added"] == 0
    assert body["failed"] == 2
    assert {f["reason"] for f in body["failures"]} == {"Invalid price"}
    assert db.read("restaurants/r1/items") is None  # nothing persisted


def test_import_missing_name_fails(client, auth_headers):
    wb = _wb_from_rows([
        ["", "", 10, "Food", "S", "veg"],
    ])
    body = _post(client, wb, auth_headers).get_json()
    assert body["added"] == 0
    assert body["failed"] == 1
    assert body["failures"][0]["reason"] == "Missing name"


def test_import_missing_category_fails(client, auth_headers):
    wb = _wb_from_rows([
        ["A", "", 10, "Food", "", "veg"],   # blank sub category
    ])
    body = _post(client, wb, auth_headers).get_json()
    assert body["failed"] == 1
    assert body["failures"][0]["reason"] == "Missing category"


def test_import_no_file(client, auth_headers):
    resp = client.post("/api/admin/menu/import",
                       data={}, content_type="multipart/form-data",
                       headers=auth_headers)
    assert resp.status_code == 400
    assert "No file" in resp.get_json()["error"]


def test_import_bad_headers(client, auth_headers):
    wb = _wb_from_rows([["Foo", "Bar"]], headers=["Foo", "Bar"])
    resp = _post(client, wb, auth_headers)
    assert resp.status_code == 400
    assert "required column" in resp.get_json()["error"]


def test_import_unreadable_file(client, auth_headers):
    resp = client.post(
        "/api/admin/menu/import",
        data={"file": (io.BytesIO(b"not an excel file"), "bad.xlsx")},
        content_type="multipart/form-data",
        headers=auth_headers,
    )
    assert resp.status_code == 400
    assert "could not read" in resp.get_json()["error"].lower()


def test_import_empty_sheet(client, auth_headers):
    wb = Workbook()  # no rows at all
    resp = _post(client, wb, auth_headers)
    assert resp.status_code == 400


# --------------------------------------------------------------------------
# Auto-create + pair resolution
# --------------------------------------------------------------------------
def test_import_autocreates_vocab_once(client, db, auth_headers):
    wb = _wb_from_rows([
        ["A", "", 10, "Food", "Cakes", "veg", "Cocoa", "French", "sweet, rich"],
        ["B", "", 10, "Food", "Cakes", "veg", "Cocoa", "French", "sweet"],
    ])
    body = _post(client, wb, auth_headers).get_json()
    assert body["added"] == 2

    assert len(db.read("restaurants/r1/main_categories")) == 1
    assert len(db.read("restaurants/r1/categories")) == 1
    assert len(db.read("restaurants/r1/ingredients")) == 1
    assert len(db.read("restaurants/r1/cuisines")) == 1
    tastes = db.read("restaurants/r1/tastes")
    assert {v["name"] for v in tastes.values()} == {"sweet", "rich"}
    # item stores names, multi-value taste as list
    items = db.read("restaurants/r1/items")
    a = next(i for i in items.values() if i["name"] == "A")
    assert a["taste"] == ["sweet", "rich"]
    assert a["main_ingredient"] == "Cocoa"


def test_import_pair_resolution_distinct_subs(client, db, auth_headers):
    wb = _wb_from_rows([
        ["A", "", 10, "Food", "Special", "veg"],
        ["B", "", 10, "Drinks", "Special", "veg"],
    ])
    _post(client, wb, auth_headers)
    cats = db.read("restaurants/r1/categories")
    specials = [c for c in cats.values() if c["name"] == "Special"]
    assert len(specials) == 2
    # two different main categories
    mains = db.read("restaurants/r1/main_categories")
    assert {m["name"] for m in mains.values()} == {"Food", "Drinks"}


# --------------------------------------------------------------------------
# Lenient coercion
# --------------------------------------------------------------------------
def test_import_lenient_coercion(client, db, auth_headers):
    wb = _wb_from_rows([
        ["A", "", 10, "Food", "S", "banana", "", "", "", "weird", "top", "maybe", "sure", ""],
    ])
    body = _post(client, wb, auth_headers).get_json()
    assert body["added"] == 1
    item = next(i for i in db.read("restaurants/r1/items").values())
    assert item["item_type"] == "veg"        # invalid -> default
    assert item["heaviness"] == "medium"     # invalid -> default
    assert item["priority"] == "medium"      # invalid -> default
    assert item["is_enabled"] is True        # "maybe" unrecognized -> default True
    assert item["is_bestseller"] is False    # "sure" unrecognized -> default False


def test_import_yes_no_bools(client, db, auth_headers):
    wb = _wb_from_rows([
        ["A", "", 10, "Food", "S", "veg", "", "", "", "light", "low", "no", "yes", ""],
    ])
    _post(client, wb, auth_headers)
    item = next(i for i in db.read("restaurants/r1/items").values())
    assert item["is_enabled"] is False
    assert item["is_bestseller"] is True


def test_import_skips_blank_rows(client, auth_headers):
    wb = _wb_from_rows([
        ["A", "", 10, "Food", "S", "veg"],
        ["", "", "", "", "", ""],          # fully blank -> skipped
        ["B", "", 12, "Food", "S", "veg"],
    ])
    body = _post(client, wb, auth_headers).get_json()
    assert body["added"] == 2
    assert body["failed"] == 0
