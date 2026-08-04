"""Excel import helpers for the menu.

Pure functions used by routes/admin.py so the logic is unit-testable without
HTTP. Two responsibilities:

1. ``build_template_workbook()`` -> openpyxl Workbook with the blank template
   (an ``Items`` sheet + a ``Read me`` sheet).
2. ``import_workbook(db_ref, restaurant_id, wb)`` -> dict report. Resolves or
   auto-creates categories + controlled vocabularies by NAME, pushes items one
   by one, and records per-row failures without aborting the rest of the file.

Storage conventions (mirrors seed_demo.py / AdminMenuManager.jsx / chat.py):
  * item.main_ingredient / item.cuisine   -> stored as the NAME (string)
  * item.taste                            -> list of NAME strings
  * item.category_id / item.main_category_id -> Firebase push-key IDs, so the
    importer must resolve human-readable category names to ids.
"""
import io
import re
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------
# Column definitions
# --------------------------------------------------------------------------
# (header label, structural-required). Structural-required means the importer
# needs the column to exist / be filled for a row to be placeable.
ITEM_COLUMNS = [
    "Name",
    "Description",
    "Price",
    "Main Category",
    "Sub Category",
    "Food Type",
    "Main Ingredient",
    "Cuisine",
    "Taste",
    "Heaviness",
    "Priority",
    "Available",
    "Bestseller",
    "Image URL",
]

# Header keys the importer requires to be present as columns. (Food Type etc.
# are optional because they coerce to defaults.)
REQUIRED_PARSE_HEADERS = ["name", "price", "main category", "sub category"]

# Enum vocabularies used for lenient coercion.
ITEM_TYPES = ("veg", "non-veg", "mixed")
HEAVINESS = ("light", "medium", "heavy")
PRIORITY = ("high", "medium", "low")
_TRUE = ("yes", "true", "1", "y", "t")
_FALSE = ("no", "false", "0", "n", "f")


class MenuImportError(Exception):
    """Unrecoverable file-level problem (-> HTTP 400, no report)."""


class _RowError(Exception):
    """Per-row problem handled inside the row try/except."""


# --------------------------------------------------------------------------
# Small value helpers
# --------------------------------------------------------------------------
def _norm(s):
    return str(s).strip().lower() if s is not None else ""


def _is_blank(v):
    return v is None or str(v).strip() == ""


def _str_or_empty(v):
    return str(v).strip() if v is not None else ""


def _parse_price(v):
    """Return a non-negative number (int when whole) or None if invalid/missing."""
    if _is_blank(v):
        return None
    try:
        p = float(v)
    except (TypeError, ValueError):
        return None
    if p != p or p < 0:  # NaN or negative -> invalid
        return None
    return int(p) if float(p).is_integer() else p


def _coerce_enum(v, allowed, default):
    key = _norm(v)
    return key if key in allowed else default


def _coerce_bool(v, default):
    if _is_blank(v):
        return default
    key = _norm(v)
    if key in _TRUE:
        return True
    if key in _FALSE:
        return False
    return default


# --------------------------------------------------------------------------
# Resolver: resolves names -> ids, auto-creating + caching within one import
# --------------------------------------------------------------------------
class _Importer:
    def __init__(self, db_ref, restaurant_id):
        self.db = db_ref
        self.base = f"restaurants/{restaurant_id}"
        self.main_cache = {}            # name_lower -> id
        self.sub_cache = {}             # (main_id, name_lower) -> id
        self.ing_cache = {}             # name_lower -> name
        self.cui_cache = {}             # name_lower -> name
        self.taste_cache = {}           # name_lower -> name
        self.used_names = set()         # name_lower already taken
        self._preload()

    @staticmethod
    def _max_order(d):
        vals = [(v.get("display_order") or 0) for v in (d or {}).values() if v]
        return max(vals) if vals else 0

    def _preload(self):
        items = self.db.child(f"{self.base}/items").get() or {}
        self.used_names = {
            _norm(v.get("name")) for v in items.values() if v and v.get("name")
        }

        main_cats = self.db.child(f"{self.base}/main_categories").get() or {}
        for cid, v in main_cats.items():
            if v and v.get("name"):
                self.main_cache[_norm(v["name"])] = cid

        subs = self.db.child(f"{self.base}/categories").get() or {}
        for cid, v in subs.items():
            if v and v.get("name"):
                self.sub_cache[(v.get("main_category_id"), _norm(v["name"]))] = cid

        ings = self.db.child(f"{self.base}/ingredients").get() or {}
        for cid, v in ings.items():
            if v and v.get("name"):
                self.ing_cache[_norm(v["name"])] = v["name"]

        cuis = self.db.child(f"{self.base}/cuisines").get() or {}
        for cid, v in cuis.items():
            if v and v.get("name"):
                self.cui_cache[_norm(v["name"])] = v["name"]

        tastes = self.db.child(f"{self.base}/tastes").get() or {}
        for cid, v in tastes.items():
            if v and v.get("name"):
                self.taste_cache[_norm(v["name"])] = v["name"]

        # display_order trackers for newly created nodes
        self.main_order = self._max_order(main_cats) + 1
        self.sub_order = {}  # main_id -> max display_order currently
        for v in subs.values():
            if v:
                mid = v.get("main_category_id")
                o = v.get("display_order") or 0
                if mid not in self.sub_order or o > self.sub_order[mid]:
                    self.sub_order[mid] = o
        self.ing_order = self._max_order(ings) + 1
        self.cui_order = self._max_order(cuis) + 1
        self.taste_order = self._max_order(tastes) + 1

    def push_item(self, item):
        return self.db.child(f"{self.base}/items").push(item)

    def resolve_main(self, name):
        key = _norm(name)
        if key in self.main_cache:
            return self.main_cache[key]
        ref = self.db.child(f"{self.base}/main_categories").push({
            "name": name.strip(),
            "display_order": self.main_order,
        })
        self.main_order += 1
        self.main_cache[key] = ref.key
        return ref.key

    def resolve_sub(self, main_id, name):
        key = (main_id, _norm(name))
        if key in self.sub_cache:
            return self.sub_cache[key]
        order = (self.sub_order.get(main_id, 0) or 0) + 1
        self.sub_order[main_id] = order
        ref = self.db.child(f"{self.base}/categories").push({
            "name": name.strip(),
            "main_category_id": main_id,
            "display_order": order,
        })
        self.sub_cache[key] = ref.key
        return ref.key

    def resolve_ingredient(self, name):
        key = _norm(name)
        if key not in self.ing_cache:
            self.db.child(f"{self.base}/ingredients").push({
                "name": name.strip(),
                "display_order": self.ing_order,
            })
            self.ing_order += 1
            self.ing_cache[key] = name.strip()
        return self.ing_cache[key]

    def resolve_cuisine(self, name):
        key = _norm(name)
        if key not in self.cui_cache:
            self.db.child(f"{self.base}/cuisines").push({
                "name": name.strip(),
                "display_order": self.cui_order,
            })
            self.cui_order += 1
            self.cui_cache[key] = name.strip()
        return self.cui_cache[key]

    def resolve_tastes(self, raw):
        if _is_blank(raw):
            return []
        out = []
        for part in re.split(r"[,/;]", str(raw)):
            name = part.strip()
            if not name:
                continue
            key = _norm(name)
            if key not in self.taste_cache:
                self.db.child(f"{self.base}/tastes").push({
                    "name": name,
                    "display_order": self.taste_order,
                    "emoji": "",
                })
                self.taste_order += 1
                self.taste_cache[key] = name
            out.append(self.taste_cache[key])
        return out


def _validate_row(name, price_raw, used_names):
    """Return a failure reason string, or None when the row is admissible."""
    if not name:
        return "Missing name"
    if _norm(name) in used_names:
        return "Duplicate name"
    if _parse_price(price_raw) is None:
        return "Invalid price"
    return None


# --------------------------------------------------------------------------
# Public API
# --------------------------------------------------------------------------
def build_template_workbook():
    """Build the blank template Workbook (Items sheet + Read me sheet)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(ITEM_COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    widths = {
        "Name": 26, "Description": 36, "Price": 9, "Main Category": 18,
        "Sub Category": 18, "Food Type": 12, "Main Ingredient": 16,
        "Cuisine": 14, "Taste": 24, "Heaviness": 12, "Priority": 10,
        "Available": 11, "Bestseller": 12, "Image URL": 36,
    }
    for col_idx, header in enumerate(ITEM_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 14)

    rm = wb.create_sheet("Read me")
    guide = [
        ("Menu Import — Quick Guide", True),
        ("", False),
        ("1. Fill the 'Items' sheet — one item per row, starting at row 2.", False),
        ("2. Required columns: Name, Price, Main Category, Sub Category.", False),
        ("3. Main/Sub Categories, Main Ingredient, Cuisine and Taste are", False),
        ("    AUTO-CREATED if the name you type does not already exist.", False),
        ("4. Name must be unique (case-insensitive). A duplicate row fails.", False),
        ("5. Rows that fail never block the rest — you get a full report.", False),
        ("", False),
        ("Allowed values:", True),
        ("   Food Type:  veg  |  non-veg  |  mixed", False),
        ("   Heaviness:  light  |  medium  |  heavy", False),
        ("   Priority:   high  |  medium  |  low", False),
        ("   Available / Bestseller:  yes  |  no", False),
        ("   Taste:  one or more, separated by commas (e.g. spicy, creamy)", False),
        ("", False),
        ("Notes:", True),
        ("   - Blank optional fields use safe defaults (e.g. Heaviness = medium).", False),
        ("   - Spice level is not imported (defaults to 0).", False),
        ("   - Image URL is optional; leave blank if you don't have one.", False),
    ]
    for i, (text, bold) in enumerate(guide, start=1):
        cell = rm.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=13)
        else:
            cell.font = Font(size=11)
    rm.column_dimensions["A"].width = 74

    return wb


# --------------------------------------------------------------------------
# Download package (template + instructions guide, zipped together)
# --------------------------------------------------------------------------
TEMPLATE_XLSX_NAME = "menu-import-template.xlsx"
GUIDE_FILENAME = "How to Import - Read me.txt"


def build_guide_text():
    """Plain-text instruction guide bundled alongside the template."""
    return (
        "MENU IMPORT - STEP-BY-STEP GUIDE\n"
        "==================================\n\n"
        "This package contains two files:\n"
        "  1. menu-import-template.xlsx  -> the Excel sheet you fill in\n"
        "  2. this file                  -> instructions\n\n"
        "---------------------------------\n"
        "HOW TO IMPORT YOUR MENU\n"
        "---------------------------------\n"
        "1. Open \"menu-import-template.xlsx\" in Excel / Google Sheets / LibreOffice.\n"
        "2. Fill in one item per row, starting at row 2. Leave row 1 (headers) as-is.\n"
        "3. Save the file (keep it as .xlsx).\n"
        "4. In RestoBot: open Menu Manager -> click \"Import Excel\" -> choose your\n"
        "   file -> Import.\n\n"
        "---------------------------------\n"
        "THE COLUMNS\n"
        "---------------------------------\n"
        "Required (must be filled):\n"
        "  Name            Item name. MUST be unique (no two items share a name).\n"
        "  Price           Number only, e.g. 250. No currency symbol.\n"
        "  Main Category   Top-level group, e.g. \"Food\", \"Beverages\".\n"
        "  Sub Category    Group under the main category, e.g. \"Curries\", \"Coffee\".\n\n"
        "Optional (left blank = safe default):\n"
        "  Description     Short text shown to customers.\n"
        "  Food Type       veg | non-veg | mixed            (blank -> veg)\n"
        "  Main Ingredient Used by the AI chat, e.g. \"Chicken\".\n"
        "  Cuisine         e.g. \"Indian\".                   (blank -> none)\n"
        "  Taste           One or more, comma-separated, e.g. \"spicy, creamy\"\n"
        "  Heaviness       light | medium | heavy           (blank -> medium)\n"
        "  Priority        high | medium | low              (blank -> medium)\n"
        "  Available       yes | no                         (blank -> yes)\n"
        "  Bestseller      yes | no                         (blank -> no)\n"
        "  Image URL       Link to a product image.\n\n"
        "---------------------------------\n"
        "SMART AUTO-CREATION\n"
        "---------------------------------\n"
        "You do NOT need to create categories / ingredients / cuisines / tastes\n"
        "beforehand. If a name you type does not already exist, RestoBot creates\n"
        "it automatically:\n"
        "  - New Main Categories and Sub Categories\n"
        "  - New Main Ingredients, Cuisines, and Tastes\n"
        "Just type the names you want and they appear.\n\n"
        "---------------------------------\n"
        "WHAT HAPPENS ON CONFLICTS\n"
        "---------------------------------\n"
        "  - Two items with the same Name? The duplicate row is skipped.\n"
        "  - Missing Name or invalid Price? That row is skipped.\n"
        "  - Skipped rows NEVER stop the rest - every other row still imports.\n"
        "  - When import finishes you get a report:\n"
        "      Added:  how many items were created\n"
        "      Failed: how many rows were skipped (with the name + reason)\n\n"
        "---------------------------------\n"
        "EXAMPLE ROW\n"
        "---------------------------------\n"
        "  Name         Price  Main Category  Sub Category  Food Type  Taste      Heaviness\n"
        "  Masala Chai  25     Beverages      Tea           veg        sweet      light\n\n"
        "---------------------------------\n"
        "TIPS\n"
        "---------------------------------\n"
        "  - Names are case-insensitive for uniqueness: \"Biryani\" and \"BIRYANI\" clash.\n"
        "  - Spice level is not imported (defaults to 0). Edit it per item later.\n"
        "  - You can re-upload this whole .zip if you like - only the .xlsx inside\n"
        "    is read.\n"
        "  - Need a fresh template? Download it again any time from Menu Manager.\n"
    )


def build_import_package():
    """Return a BytesIO ZIP containing the blank template + the instruction guide."""
    wb = build_template_workbook()
    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_bytes = xlsx_buf.getvalue()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(TEMPLATE_XLSX_NAME, xlsx_bytes)
        zf.writestr(GUIDE_FILENAME, build_guide_text())
    buf.seek(0)
    return buf


def workbook_from_upload(file_storage):
    """Read an uploaded file and return an openpyxl Workbook.

    Accepts a plain .xlsx, OR a .zip that contains an .xlsx (so a user may
    re-upload the downloaded package as-is). Raises MenuImportError otherwise.
    """
    raw = file_storage.read()
    if not raw:
        raise MenuImportError("The uploaded file is empty.")

    # 1) Plain xlsx (an xlsx is itself a zip, so try openpyxl first).
    try:
        return load_workbook(io.BytesIO(raw), data_only=True)
    except Exception:
        pass

    # 2) A wrapper zip that holds the .xlsx (e.g. our download package).
    try:
        if zipfile.is_zipfile(io.BytesIO(raw)):
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                names = [
                    n for n in zf.namelist()
                    if n.lower().endswith(".xlsx") and not n.startswith("__MACOSX")
                ]
                if names:
                    return load_workbook(io.BytesIO(zf.read(names[0])), data_only=True)
    except Exception:
        pass

    raise MenuImportError("not-a-valid-workbook")


def import_workbook(db_ref, restaurant_id, wb):
    """Import items from a workbook. Returns a report dict.

    Never raises for per-row problems (those become report entries). Raises
    MenuImportError only for unrecoverable file-level issues (empty sheet,
    missing required columns).
    """
    sheet = wb["Items"] if "Items" in wb.sheetnames else wb.active
    all_rows = list(sheet.iter_rows(values_only=True))

    # Find the header = first non-empty row.
    header = None
    data_start = 0
    for i, r in enumerate(all_rows):
        if any(not _is_blank(c) for c in r):
            header = r
            data_start = i + 1
            break
    if header is None:
        raise MenuImportError("The sheet is empty.")

    col_map = {}
    for idx, h in enumerate(header):
        key = _norm(h)
        if key and key not in col_map:
            col_map[key] = idx

    missing = [h for h in REQUIRED_PARSE_HEADERS if h not in col_map]
    if missing:
        raise MenuImportError(
            "Missing required column(s): " + ", ".join(missing) +
            ". Download the template to get the correct structure."
        )

    imp = _Importer(db_ref, restaurant_id)

    def cell(row, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    added = 0
    failed = 0
    added_names = []
    failures = []

    for offset, row in enumerate(all_rows[data_start:], start=data_start + 1):
        if not any(not _is_blank(c) for c in row):
            continue  # skip fully blank rows

        name = _str_or_empty(cell(row, "name"))

        reason = _validate_row(name, cell(row, "price"), imp.used_names)
        if reason:
            failed += 1
            failures.append({"row": offset, "name": name, "reason": reason})
            continue

        try:
            main_name = _str_or_empty(cell(row, "main category"))
            sub_name = _str_or_empty(cell(row, "sub category"))
            if not main_name or not sub_name:
                raise _RowError("Missing category")

            main_id = imp.resolve_main(main_name)
            sub_id = imp.resolve_sub(main_id, sub_name)

            item = {
                "name": name,
                "description": _str_or_empty(cell(row, "description")),
                "price": _parse_price(cell(row, "price")),
                "main_category_id": main_id,
                "category_id": sub_id,
                "item_type": _coerce_enum(cell(row, "food type"), ITEM_TYPES, "veg"),
                "main_ingredient": _resolve_optional(cell(row, "main ingredient"), imp.resolve_ingredient),
                "cuisine": _resolve_optional(cell(row, "cuisine"), imp.resolve_cuisine),
                "taste": imp.resolve_tastes(cell(row, "taste")),
                "heaviness": _coerce_enum(cell(row, "heaviness"), HEAVINESS, "medium"),
                "priority": _coerce_enum(cell(row, "priority"), PRIORITY, "medium"),
                "is_enabled": _coerce_bool(cell(row, "available"), True),
                "is_bestseller": _coerce_bool(cell(row, "bestseller"), False),
                "image_url": _str_or_empty(cell(row, "image url")),
                "spice_level": 0,
            }
            imp.push_item(item)
            imp.used_names.add(_norm(name))
            added += 1
            added_names.append(name)
        except _RowError as e:
            failed += 1
            failures.append({"row": offset, "name": name, "reason": str(e)})
        except Exception:
            # Defensive: never abort the whole file on an unexpected error.
            failed += 1
            failures.append({"row": offset, "name": name, "reason": "Import error"})

    return {
        "added": added,
        "failed": failed,
        "total": added + failed,
        "added_names": added_names,
        "failures": failures,
    }


def _resolve_optional(v, resolver):
    if _is_blank(v):
        return ""
    return resolver(str(v).strip())
